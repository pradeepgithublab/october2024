import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# Replace with your Mist API credentials
MIST_API_TOKEN = 'xh9oBdbYQllXCmCh5Ad63h3Y3enrewPWrDD37XpYA5f1ealYTPJQEb2FMkmo0DBi9vFIiLJpNlHd6fM95Zog8e0NldKTnOol'
ORG_ID = '15e7597e-9b06-4381-8443-16aba95c5e0d'
SITE_ID = 'a631d811-8b9b-48ef-823b-b5594e4af356'

# Function to get live users and filter by signal strength
def get_live_users():
    headers = {
        'Authorization': f'Token {MIST_API_TOKEN}'
    }
    url = f'https://api.eu.mist.com/api/v1/sites/{SITE_ID}/stats/clients'
    response = requests.get(url, headers=headers)
    
    if response.status_code == 200:
        clients = response.json()
        
        # Filter clients based on signal strength (RSSI)
        filtered_clients = [
            {
                'MAC Address': client.get('mac'),
                'SSID': client.get('ssid', 'N/A'),
                'Signal Strength (dBm)': client.get('rssi', -100)
            }
            for client in clients
            if client.get('rssi', -100) >= -70  # only include those with stronger signals
        ]
        
        return filtered_clients
    else:
        print(f"Error: {response.status_code}, {response.text}")
        return []

# Function to save data to Excel and apply formatting
def save_to_excel_with_formatting(clients_data):
    # Convert to DataFrame
    df = pd.DataFrame(clients_data)
    
    # Save DataFrame to Excel
    file_path = 'Client_Signal_Strength.xlsx'
    df.to_excel(file_path, index=False, sheet_name='Client Data')
    
    # Load workbook and apply formatting for signal strength
    workbook = load_workbook(file_path)
    sheet = workbook['Client Data']
    
    # Define red font for low signal strength
    red_font = Font(color="FF0000")  # Red for signal strength below -70 dBm
    
    # Apply conditional formatting
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=3, max_col=3):
        cell = row[0]
        if cell.value is not None and cell.value < -70:
            cell.font = red_font

    # Save the updated workbook
    workbook.save(file_path)
    print("Data has been saved with color formatting to Client_Signal_Strength.xlsx")

if __name__ == "__main__":
    clients_data = get_live_users()
    if clients_data:
        save_to_excel_with_formatting(clients_data)
    else:
        print("No data to save.")
