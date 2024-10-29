import requests
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

# Replace with your Juniper Mist API credentials and endpoint
api_token = 'xh9oBdbYQllXCmCh5Ad63h3Y3enrewPWrDD37XpYA5f1ealYTPJQEb2FMkmo0DBi9vFIiLJpNlHd6fM95Zog8e0NldKTnOol'
org_id = '15e7597e-9b06-4381-8443-16aba95c5e0d'
base_url = 'https://api.eu.mist.com/api/v1'

# Headers for requests
headers = {
    'Content-Type': 'application/json',
    'Authorization': f'Token {api_token}',
}

# Function to get sites
def get_sites():
    try:
        response = requests.get(f'{base_url}/orgs/{org_id}/sites', headers=headers)
        response.raise_for_status()
        return response.json()
    except requests.RequestException as e:
        print(f"Error fetching sites: {e}")
        return []

# Function to get AP data for a specific site
def get_site_ap_data(site):
    try:
        site_id = site['id']
        site_name = site['name']
        
        # Fetch device stats for this site
        response = requests.get(f'{base_url}/sites/{site_id}/stats/devices', headers=headers)
        response.raise_for_status()
        
        devices = response.json()
        connected = len([ap for ap in devices if ap['status'] == 'connected'])
        disconnected = len([ap for ap in devices if ap['status'] == 'disconnected'])
        
        return {'Location': site_name, 'Connected APs': connected, 'Disconnected APs': disconnected}
    except requests.RequestException as e:
        print(f"Error fetching data for site {site['name']}: {e}")
        return {'Location': site['name'], 'Connected APs': None, 'Disconnected APs': None}

# Function to collect all AP data in parallel
def collect_ap_data_parallel(sites):
    ap_data = []
    with ThreadPoolExecutor(max_workers=5) as executor:
        future_to_site = {executor.submit(get_site_ap_data, site): site for site in sites}
        for future in as_completed(future_to_site):
            result = future.result()
            if result:
                ap_data.append(result)
    return ap_data

# Function to save data to Excel with colored formatting
def save_to_excel_with_formatting(ap_data):
    # Convert data to DataFrame
    df = pd.DataFrame(ap_data)
    
    # Save DataFrame to Excel temporarily
    file_path = 'AP_Connection_Status.xlsx'
    df.to_excel(file_path, index=False)
    
    # Load the workbook and access the sheet
    workbook = load_workbook(file_path)
    sheet = workbook.active
    
    # Define font styles
    dark_green_font = Font(color="008000")  # Dark green for connected APs
    red_font = Font(color="FF0000")         # Red for disconnected APs
    
    # Apply conditional formatting
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3):
        site_name_cell, connected_cell, disconnected_cell = row
        
        # Apply dark green font for connected APs and their site name if there are connected APs
        if connected_cell.value and connected_cell.value > 0:
            site_name_cell.font = dark_green_font
            connected_cell.font = dark_green_font
        
        # Apply red font for disconnected APs and their site name if there are disconnected APs
        if disconnected_cell.value and disconnected_cell.value > 0:
            site_name_cell.font = red_font
            disconnected_cell.font = red_font

    # Save the updated workbook with formatting
    workbook.save(file_path)
    print("Data has been saved with color formatting to AP_Connection_Status.xlsx")

if __name__ == '__main__':
    start_time = time.time()
    sites = get_sites()
    if sites:
        ap_data = collect_ap_data_parallel(sites)
        save_to_excel_with_formatting(ap_data)
    else:
        print("No sites found or error in fetching sites.")
    print(f"Execution time: {time.time() - start_time:.2f} seconds")
